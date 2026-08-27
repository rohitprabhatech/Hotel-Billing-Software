"""Owner sales reporting and exports."""

import calendar
import csv
import io
import re
from datetime import datetime

from flask import current_app, send_file
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

from app.constants.payments import normalize_payment_method, payment_method_label
from app.constants.permissions import PERM_REPORTS
from app.constants.report_registry import (
    DEFAULT_BILLS_PER_PAGE,
    MAX_BILLS_PER_PAGE,
    MAX_CUSTOM_RANGE_DAYS,
    filter_registry_for_modules,
)
from app.extensions import db
from app.repositories.report_repository import ReportRepository
from app.repositories.tenant_repository import TenantRepository
from app.services.audit_service import AuditService
from app.services.module_service import ModuleService
from app.utils.exceptions import NotFoundError, ValidationError
from app.utils.permission_access import require_permission
from app.utils.periods import fill_day_wise_series, resolve_period
from app.utils.request_context import require_request_context

METRIC_LABELS = {
    "total_sales": "Total Sales",
    "bill_count": "Bills",
    "total_discount": "Discount",
    "total_gst": "GST",
    "average_bill": "Average Bill",
    "items_sold": "Items Sold",
    "cancelled_bills": "Cancelled Bills",
    "cash_sales": "Cash Sales",
    "online_sales": "Online Sales",
    "credit_sales": "Credit Sales",
    "cash_bill_count": "Cash Bills",
    "online_bill_count": "Online Bills",
    "credit_bill_count": "Credit Bills",
}

TOP_ITEMS_LIMIT = 5


class ReportService:
    @staticmethod
    def _ensure_reports_access():
        require_permission(PERM_REPORTS)
        return require_request_context()

    @staticmethod
    def _tz():
        return current_app.config.get("REPORT_TIMEZONE", "Asia/Kolkata")

    @staticmethod
    def _bounds(period: str, from_date=None, to_date=None):
        try:
            return resolve_period(period, ReportService._tz(), from_date, to_date)
        except ValueError as exc:
            raise ValidationError(str(exc)) from exc

    @staticmethod
    def _normalize_filter(payment_method: str | None):
        if not payment_method:
            return None
        try:
            return normalize_payment_method(payment_method)
        except ValueError as exc:
            raise ValidationError("payment_method must be cash, online, or credit") from exc

    @staticmethod
    def _serialize_bills(bills):
        return [
            {
                "id": bill.id,
                "bill_number": bill.bill_number,
                "created_at": bill.created_at.isoformat() if bill.created_at else None,
                "grand_total": float(bill.grand_total),
                "status": bill.status,
                "payment_method": bill.payment_method,
                "payment_method_label": payment_method_label(bill.payment_method),
            }
            for bill in bills
        ]

    @staticmethod
    def _rank_items(item_wise: list[dict]):
        top_items = item_wise[:TOP_ITEMS_LIMIT]
        low_items = sorted(item_wise, key=lambda row: (row["revenue"], row["quantity"]))[
            :TOP_ITEMS_LIMIT
        ]
        return top_items, low_items

    @staticmethod
    def _bills_page(page=None, per_page=None) -> tuple[int, int]:
        page_n = max(int(page or 1), 1)
        per_n = min(
            max(int(per_page or DEFAULT_BILLS_PER_PAGE), 1),
            MAX_BILLS_PER_PAGE,
        )
        return page_n, per_n

    @staticmethod
    def available_reports():
        ctx = ReportService._ensure_reports_access()
        tenant = TenantRepository.get_by_id(ctx.tenant_id)
        if tenant is None:
            raise NotFoundError("Tenant not found")
        enabled = set(ModuleService.enabled_codes_for_tenant(tenant))
        reports = filter_registry_for_modules(enabled)
        return {
            "reports": reports,
            "hub_reports": [row for row in reports if row["kind"] == "hub"],
            "link_reports": [row for row in reports if row["kind"] == "link"],
            "limits": {
                "max_custom_range_days": MAX_CUSTOM_RANGE_DAYS,
                "default_bills_per_page": DEFAULT_BILLS_PER_PAGE,
                "max_bills_per_page": MAX_BILLS_PER_PAGE,
            },
            "index_notes": [
                "bills: ix_bills_tenant_status_created_at (tenant_id, status, created_at)",
            ],
        }

    @staticmethod
    def _build_report(
        tenant_id: str,
        start,
        end,
        label: str,
        period: str,
        payment_method: str | None = None,
        *,
        page: int = 1,
        per_page: int = DEFAULT_BILLS_PER_PAGE,
    ):
        method = ReportService._normalize_filter(payment_method)
        page_n, per_n = ReportService._bills_page(page, per_page)
        item_wise = ReportRepository.item_wise(
            tenant_id, start, end, payment_method=method
        )
        top_items, low_items = ReportService._rank_items(item_wise)
        bills, bills_total = ReportRepository.bill_rows(
            tenant_id,
            start,
            end,
            payment_method=method,
            page=page_n,
            per_page=per_n,
        )
        return {
            "period": period,
            "label": label,
            "payment_method": method,
            "metrics": ReportRepository.period_metrics(
                tenant_id, start, end, payment_method=method
            ),
            "item_wise": item_wise,
            "top_items": top_items,
            "low_items": low_items,
            "category_wise": ReportRepository.category_wise(
                tenant_id, start, end, payment_method=method
            ),
            "day_wise": fill_day_wise_series(
                ReportRepository.day_wise(
                    tenant_id,
                    start,
                    end,
                    payment_method=method,
                    tz_name=ReportService._tz(),
                ),
                start,
                end,
                ReportService._tz(),
            ),
            "bills": ReportService._serialize_bills(bills),
            "bills_meta": {
                "page": page_n,
                "per_page": per_n,
                "total": bills_total,
            },
        }

    @staticmethod
    def summary(period: str = "today", from_date=None, to_date=None):
        ctx = ReportService._ensure_reports_access()
        start, end, label, prev_start, prev_end, prev_label = ReportService._bounds(
            period, from_date, to_date
        )
        current = ReportRepository.period_metrics(ctx.tenant_id, start, end)
        previous = ReportRepository.period_metrics(ctx.tenant_id, prev_start, prev_end)
        item_wise = ReportRepository.item_wise(ctx.tenant_id, start, end)
        top_items, low_items = ReportService._rank_items(item_wise)
        from app.repositories.bill_delivery_repository import BillDeliveryRepository

        whatsapp_delivery = BillDeliveryRepository.whatsapp_status_counts(
            ctx.tenant_id, date_from=start, date_to=end
        )
        email_delivery = BillDeliveryRepository.email_status_counts(
            ctx.tenant_id, date_from=start, date_to=end
        )
        from app.repositories.item_repository import ItemRepository

        inventory_health = ItemRepository.inventory_health_counts(ctx.tenant_id)
        tz_name = ReportService._tz()
        day_wise = ReportRepository.day_wise(
            ctx.tenant_id, start, end, tz_name=tz_name
        )
        day_wise = fill_day_wise_series(day_wise, start, end, tz_name)
        return {
            "period": period,
            "label": label,
            "previous_label": prev_label,
            "current": current,
            "previous": previous,
            "item_wise": item_wise[:10],
            "top_items": top_items,
            "low_items": low_items,
            "category_wise": ReportRepository.category_wise(ctx.tenant_id, start, end),
            "day_wise": day_wise,
            "whatsapp_delivery": whatsapp_delivery,
            "email_delivery": email_delivery,
            "inventory_health": inventory_health,
        }

    @staticmethod
    def daily_sales(
        date: str | None = None,
        payment_method: str | None = None,
        *,
        page: int = 1,
        per_page: int = DEFAULT_BILLS_PER_PAGE,
    ):
        ctx = ReportService._ensure_reports_access()
        if date:
            start, end, label, *_ = ReportService._bounds("custom", date, date)
        else:
            start, end, label, *_ = ReportService._bounds("today")
        return ReportService._build_report(
            ctx.tenant_id,
            start,
            end,
            label,
            "daily",
            payment_method=payment_method,
            page=page,
            per_page=per_page,
        )

    @staticmethod
    def weekly_sales(
        payment_method: str | None = None,
        *,
        page: int = 1,
        per_page: int = DEFAULT_BILLS_PER_PAGE,
    ):
        ctx = ReportService._ensure_reports_access()
        start, end, label, *_ = ReportService._bounds("this_week")
        return ReportService._build_report(
            ctx.tenant_id,
            start,
            end,
            label,
            "weekly",
            payment_method=payment_method,
            page=page,
            per_page=per_page,
        )

    @staticmethod
    def monthly_sales(
        year: int | None = None,
        month: int | None = None,
        payment_method: str | None = None,
        *,
        page: int = 1,
        per_page: int = DEFAULT_BILLS_PER_PAGE,
    ):
        ctx = ReportService._ensure_reports_access()
        if year and month:
            last = calendar.monthrange(int(year), int(month))[1]
            from_date = f"{int(year):04d}-{int(month):02d}-01"
            to_date = f"{int(year):04d}-{int(month):02d}-{last:02d}"
            start, end, label, *_ = ReportService._bounds("custom", from_date, to_date)
        else:
            start, end, label, *_ = ReportService._bounds("this_month")
        return ReportService._build_report(
            ctx.tenant_id,
            start,
            end,
            label,
            "monthly",
            payment_method=payment_method,
            page=page,
            per_page=per_page,
        )

    @staticmethod
    def custom_sales(
        from_date: str,
        to_date: str,
        payment_method: str | None = None,
        *,
        page: int = 1,
        per_page: int = DEFAULT_BILLS_PER_PAGE,
    ):
        ctx = ReportService._ensure_reports_access()
        start, end, label, *_ = ReportService._bounds("custom", from_date, to_date)
        return ReportService._build_report(
            ctx.tenant_id,
            start,
            end,
            label,
            "custom",
            payment_method=payment_method,
            page=page,
            per_page=per_page,
        )

    @staticmethod
    def fb_report(*, date: str | None = None, from_date: str | None = None, to_date: str | None = None):
        ctx = ReportService._ensure_reports_access()
        if from_date and to_date:
            start, end, label, *_ = ReportService._bounds("custom", from_date, to_date)
            period = "custom"
        elif date:
            start, end, label, *_ = ReportService._bounds("custom", date, date)
            period = "daily"
        else:
            start, end, label, *_ = ReportService._bounds("today")
            period = "daily"

        from app.repositories.fb_report_repository import FbReportRepository

        metrics = ReportRepository.period_metrics(ctx.tenant_id, start, end)
        wastage = FbReportRepository.wastage_summary(ctx.tenant_id, start, end)
        return {
            "period": period,
            "label": label,
            "metrics": metrics,
            "channel_wise": FbReportRepository.channel_wise(ctx.tenant_id, start, end),
            "table_wise": FbReportRepository.table_wise(ctx.tenant_id, start, end),
            "wastage": wastage,
        }

    @staticmethod
    def export(
        *,
        report_type: str,
        fmt: str,
        date=None,
        from_date=None,
        to_date=None,
        year=None,
        month=None,
        payment_method=None,
    ):
        ctx = ReportService._ensure_reports_access()
        fmt = (fmt or "xlsx").lower()
        if fmt not in {"xlsx", "csv", "pdf"}:
            raise ValidationError("format must be xlsx, csv, or pdf")

        report_type = (report_type or "daily").lower()
        if report_type == "daily":
            report = ReportService.daily_sales(date, payment_method=payment_method)
        elif report_type == "weekly":
            report = ReportService.weekly_sales(payment_method=payment_method)
        elif report_type == "monthly":
            report = ReportService.monthly_sales(
                int(year) if year else None,
                int(month) if month else None,
                payment_method=payment_method,
            )
        elif report_type == "custom":
            report = ReportService.custom_sales(
                from_date, to_date, payment_method=payment_method
            )
        else:
            raise ValidationError("type must be daily, weekly, monthly, or custom")

        tenant = TenantRepository.get_by_id(ctx.tenant_id)
        business = (tenant.business_name if tenant else "Business").strip() or "Business"
        safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", business)
        safe_period = re.sub(r"[^A-Za-z0-9_-]+", "_", report["label"])
        filename = f"{safe_name}_{safe_period}_Sales.{fmt}"

        AuditService.log(
            tenant_id=ctx.tenant_id,
            action="EXPORT_REPORT",
            entity_type="REPORT",
            entity_id=None,
            new_data={
                "type": report_type,
                "label": report["label"],
                "format": fmt,
                "filename": filename,
            },
        )
        db.session.commit()

        if fmt == "csv":
            return ReportService._export_csv(report, filename)
        if fmt == "pdf":
            return ReportService._export_pdf(report, filename, business)
        return ReportService._export_xlsx(report, filename)

    @staticmethod
    def _metric_rows(metrics: dict):
        return [
            (METRIC_LABELS.get(key, key.replace("_", " ").title()), value)
            for key, value in metrics.items()
        ]

    @staticmethod
    def _export_csv(report, filename):
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(["Metric", "Value"])
        for label, value in ReportService._metric_rows(report["metrics"]):
            writer.writerow([label, value])
        writer.writerow([])
        writer.writerow(["Top Items"])
        writer.writerow(["Item", "Quantity", "Revenue"])
        for row in report.get("top_items") or []:
            writer.writerow([row["item_name"], row["quantity"], row["revenue"]])
        writer.writerow([])
        writer.writerow(["Low Items"])
        writer.writerow(["Item", "Quantity", "Revenue"])
        for row in report.get("low_items") or []:
            writer.writerow([row["item_name"], row["quantity"], row["revenue"]])
        writer.writerow([])
        writer.writerow(["Category", "Quantity", "Revenue"])
        for row in report.get("category_wise") or []:
            writer.writerow([row["category_name"], row["quantity"], row["revenue"]])
        writer.writerow([])
        writer.writerow(["Item", "Quantity", "Revenue"])
        for row in report["item_wise"]:
            writer.writerow([row["item_name"], row["quantity"], row["revenue"]])
        writer.writerow([])
        writer.writerow(["Date", "Sales", "Bills"])
        for row in report["day_wise"]:
            writer.writerow([row["date"], row["total_sales"], row["bill_count"]])
        writer.writerow([])
        writer.writerow(["Bill No", "Date", "Amount", "Payment Method", "Status"])
        for row in report.get("bills") or []:
            writer.writerow(
                [
                    row["bill_number"],
                    row["created_at"],
                    row["grand_total"],
                    row["payment_method_label"],
                    row["status"],
                ]
            )

        mem = io.BytesIO(buffer.getvalue().encode("utf-8"))
        mem.seek(0)
        return send_file(
            mem,
            mimetype="text/csv",
            as_attachment=True,
            download_name=filename,
        )

    @staticmethod
    def _export_xlsx(report, filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws.append(["Metric", "Value"])
        for label, value in ReportService._metric_rows(report["metrics"]):
            ws.append([label, value])

        ws_top = wb.create_sheet("Top Items")
        ws_top.append(["Item", "Quantity", "Revenue"])
        for row in report.get("top_items") or []:
            ws_top.append([row["item_name"], row["quantity"], row["revenue"]])

        ws_low = wb.create_sheet("Low Items")
        ws_low.append(["Item", "Quantity", "Revenue"])
        for row in report.get("low_items") or []:
            ws_low.append([row["item_name"], row["quantity"], row["revenue"]])

        ws_cat = wb.create_sheet("Category Sales")
        ws_cat.append(["Category", "Quantity", "Revenue"])
        for row in report.get("category_wise") or []:
            ws_cat.append([row["category_name"], row["quantity"], row["revenue"]])

        ws2 = wb.create_sheet("Item Wise")
        ws2.append(["Item", "Quantity", "Revenue"])
        for row in report["item_wise"]:
            ws2.append([row["item_name"], row["quantity"], row["revenue"]])

        ws3 = wb.create_sheet("Day Wise")
        ws3.append(["Date", "Sales", "Bills"])
        for row in report["day_wise"]:
            ws3.append([row["date"], row["total_sales"], row["bill_count"]])

        ws4 = wb.create_sheet("Bills")
        ws4.append(["Bill No", "Date", "Amount", "Payment Method", "Status"])
        for row in report.get("bills") or []:
            ws4.append(
                [
                    row["bill_number"],
                    row["created_at"],
                    row["grand_total"],
                    row["payment_method_label"],
                    row["status"],
                ]
            )

        mem = io.BytesIO()
        wb.save(mem)
        mem.seek(0)
        return send_file(
            mem,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )

    @staticmethod
    def outstanding(*, party_type: str | None = None, as_of: str | None = None):
        """Aged customer + supplier outstanding (BIZ-54)."""
        ReportService._ensure_reports_access()
        from app.services.party_ledger_service import PartyLedgerService

        return PartyLedgerService.aged_outstanding_report(
            as_of=as_of or None,
            party_type=party_type,
        )

    @staticmethod
    def _export_pdf(report, filename, business_name):
        mem = io.BytesIO()
        c = canvas.Canvas(mem, pagesize=A4)
        _width, height = A4
        y = height - 50
        c.setFont("Helvetica-Bold", 14)
        c.drawString(40, y, f"{business_name} - Sales Report")
        y -= 24
        c.setFont("Helvetica", 11)
        c.drawString(40, y, f"Period: {report['label']}")
        y -= 20
        c.drawString(40, y, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        y -= 30
        for label, value in ReportService._metric_rows(report["metrics"]):
            c.drawString(40, y, f"{label}: {value}")
            y -= 16
            if y < 80:
                c.showPage()
                y = height - 50
                c.setFont("Helvetica", 11)
        c.showPage()
        c.save()
        mem.seek(0)
        return send_file(
            mem,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=filename,
        )
